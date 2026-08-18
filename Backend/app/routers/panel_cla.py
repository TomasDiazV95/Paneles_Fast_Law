import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from app.db.database import engine
from app.routers.auth import get_current_user
from app.schemas.panel_cla import (
    AvanceEtapaFilaCLA,
    ComparativoFilaCLA,
    ComparativoResponseCLA,
    ContactabilidadFilaCLA,
    ContactabilidadResponseCLA,
    EjecutivoCLA,
    EstadoCarteraCLA,
    InboundCLA,
    PagosDiarioCLA,
    PagosFilaCLA,
    PagosResponseCLA,
    ProductividadFilaCLA,
    ProductividadResponseCLA,
    ReprosDiarioCLA,
    ReprosFilaCLA,
    ReprosResponseCLA,
    ResumenContactoCLA,
)


def _periodo_anterior(periodo: str) -> str:
    anio, mes = int(periodo[:4]), int(periodo[4:])
    if mes == 1:
        return f"{anio - 1}12"
    return f"{anio}{mes - 1:02d}"


def _valor(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _ejecutar_multi_resultset(sql: str, params: tuple) -> dict[str, list[dict]]:
    raw_conn = engine.raw_connection()
    try:
        cursor = raw_conn.cursor()
        cursor.execute(sql, params)
        datasets: dict[str, list[dict]] = {}
        while True:
            columns = [c[0] for c in cursor.description] if cursor.description else []
            if columns == ["DATASET"]:
                nombre = cursor.fetchone()[0]
                if not cursor.nextset():
                    break
                columnas = [c[0] for c in cursor.description]
                filas = [dict(zip(columnas, (_valor(v) for v in row))) for row in cursor.fetchall()]
                datasets[nombre] = filas
            if not cursor.nextset():
                break
        return datasets
    finally:
        raw_conn.close()

router = APIRouter(prefix="/panel/cla", tags=["panel-cla"], dependencies=[Depends(get_current_user)])


@router.get("/estado-cartera", response_model=list[EstadoCarteraCLA])
def estado_cartera(periodo: str):
    query = text(
        """
        SELECT CLASIFICACION, Q_CAUSAS, CUANTIA_TOTAL, TICKET_PROMEDIO, PCT_DISTRIBUCION
        FROM dbo.PANEL1_ESTADO_CARTERA
        WHERE PERIODO = :periodo
        ORDER BY CLASIFICACION
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"periodo": periodo}).mappings().all()

    return [
        EstadoCarteraCLA(
            clasificacion=r["CLASIFICACION"],
            cantidad_causas=r["Q_CAUSAS"],
            cuantia_total=r["CUANTIA_TOTAL"],
            ticket_promedio=r["TICKET_PROMEDIO"],
            pct_distribucion=r["PCT_DISTRIBUCION"],
        )
        for r in rows
    ]


@router.get("/contactabilidad", response_model=ContactabilidadResponseCLA)
def contactabilidad(periodo: str):
    matriz_query = text(
        """
        SELECT CLASIFICACION, TIPO_CONTACTO, Q_CAUSAS, PCT_DENTRO_CLASIFICACION, PROMEDIO_INTENSIDAD, DIAS_HABILES
        FROM dbo.PANEL1_CONTACTABILIDAD
        WHERE PERIODO = :periodo
        ORDER BY CLASIFICACION, TIPO_CONTACTO
        """
    )
    resumen_query = text(
        """
        SELECT TIPO_CONTACTO, Q_CAUSAS, PCT, DIAS_HABILES, TOTAL_GESTIONES, PROMEDIO_GESTIONES_DIARIAS
        FROM dbo.PANEL1_RESUMEN_CONTACTO
        WHERE PERIODO = :periodo
        ORDER BY Q_CAUSAS DESC
        """
    )
    inbound_query = text(
        """
        SELECT CLASIFICACION, Q_TOTAL, CONTACTO_DIRECTO_INB, SIN_CONTACTO_INBOUND
        FROM dbo.PANEL1_INBOUND
        WHERE PERIODO = :periodo
        """
    )
    with engine.connect() as conn:
        matriz = conn.execute(matriz_query, {"periodo": periodo}).mappings().all()
        resumen = conn.execute(resumen_query, {"periodo": periodo}).mappings().all()
        inbound = conn.execute(inbound_query, {"periodo": periodo}).mappings().all()

    return ContactabilidadResponseCLA(
        matriz=[
            ContactabilidadFilaCLA(
                clasificacion=r["CLASIFICACION"],
                tipo_contacto=r["TIPO_CONTACTO"],
                cantidad_causas=r["Q_CAUSAS"],
                pct_dentro_clasificacion=r["PCT_DENTRO_CLASIFICACION"],
                promedio_intensidad=r["PROMEDIO_INTENSIDAD"],
                dias_habiles=r["DIAS_HABILES"],
            )
            for r in matriz
        ],
        resumen=[
            ResumenContactoCLA(
                tipo_contacto=r["TIPO_CONTACTO"],
                cantidad_causas=r["Q_CAUSAS"],
                pct=r["PCT"],
                dias_habiles=r["DIAS_HABILES"],
                total_gestiones=r["TOTAL_GESTIONES"],
                promedio_gestiones_diarias=r["PROMEDIO_GESTIONES_DIARIAS"],
            )
            for r in resumen
        ],
        inbound=[
            InboundCLA(
                clasificacion=r["CLASIFICACION"],
                q_total=r["Q_TOTAL"],
                contacto_directo_inbound=r["CONTACTO_DIRECTO_INB"],
                sin_contacto_inbound=r["SIN_CONTACTO_INBOUND"],
            )
            for r in inbound
        ],
    )


@router.get("/pagos", response_model=PagosResponseCLA)
def pagos(periodo: str):
    resumen_query = text(
        """
        SELECT CLASIFICACION, Q_CAUSAS, PCT_DISTRIBUCION, TOTAL_PAGOS, TICKET_RECUPERO
        FROM dbo.PANEL1_PAGOS
        WHERE PERIODO = :periodo
        ORDER BY CASE WHEN CLASIFICACION = 'TOTAL GENERAL' THEN 1 ELSE 0 END, TOTAL_PAGOS DESC
        """
    )
    diario_query = text(
        """
        SELECT FECHA_PAGO, MONTO_DIA, MONTO_ACUMULADO
        FROM dbo.PANEL1_PAGOS_DIARIO
        WHERE PERIODO = :periodo
        ORDER BY FECHA_PAGO
        """
    )
    with engine.connect() as conn:
        resumen = conn.execute(resumen_query, {"periodo": periodo}).mappings().all()
        diario = conn.execute(diario_query, {"periodo": periodo}).mappings().all()

    return PagosResponseCLA(
        resumen=[
            PagosFilaCLA(
                clasificacion=r["CLASIFICACION"],
                cantidad_causas=r["Q_CAUSAS"],
                pct_distribucion=r["PCT_DISTRIBUCION"],
                total_pagos=r["TOTAL_PAGOS"],
                ticket_recupero=r["TICKET_RECUPERO"],
            )
            for r in resumen
        ],
        diario=[
            PagosDiarioCLA(
                fecha_pago=r["FECHA_PAGO"],
                monto_dia=r["MONTO_DIA"],
                monto_acumulado=r["MONTO_ACUMULADO"],
            )
            for r in diario
        ],
    )


@router.get("/repros", response_model=ReprosResponseCLA)
def repros(periodo: str):
    resumen_query = text(
        """
        SELECT CLASIFICACION, Q_CAUSAS, PCT_DISTRIBUCION, TOTAL_REPRO, TICKET_RECUPERO
        FROM dbo.PANEL1_REPROS
        WHERE PERIODO = :periodo
        ORDER BY CASE WHEN CLASIFICACION = 'TOTAL GENERAL' THEN 1 ELSE 0 END, TOTAL_REPRO DESC
        """
    )
    diario_query = text(
        """
        SELECT FECHA_REPRO, SALDO_DIA, SALDO_ACUMULADO
        FROM dbo.PANEL1_REPROS_DIARIO
        WHERE PERIODO = :periodo
        ORDER BY FECHA_REPRO
        """
    )
    with engine.connect() as conn:
        resumen = conn.execute(resumen_query, {"periodo": periodo}).mappings().all()
        diario = conn.execute(diario_query, {"periodo": periodo}).mappings().all()

    return ReprosResponseCLA(
        resumen=[
            ReprosFilaCLA(
                clasificacion=r["CLASIFICACION"],
                cantidad_causas=r["Q_CAUSAS"],
                pct_distribucion=r["PCT_DISTRIBUCION"],
                total_repro=r["TOTAL_REPRO"],
                ticket_recupero=r["TICKET_RECUPERO"],
            )
            for r in resumen
        ],
        diario=[
            ReprosDiarioCLA(
                fecha_repro=r["FECHA_REPRO"],
                saldo_dia=r["SALDO_DIA"],
                saldo_acumulado=r["SALDO_ACUMULADO"],
            )
            for r in diario
        ],
    )


@router.get("/ejecutivos", response_model=list[EjecutivoCLA])
def ejecutivos(periodo: str):
    query = text(
        """
        SELECT CODIGO_USUARIO, Q_GESTIONES, Q_CONTACTOS, Q_RUT_PAGOS, MONTO_PAGOS, Q_RUT_REPROS, MONTO_REPROS
        FROM dbo.PANEL1_EJECUTIVOS
        WHERE PERIODO = :periodo
        ORDER BY MONTO_PAGOS DESC
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"periodo": periodo}).mappings().all()

    return [
        EjecutivoCLA(
            codigo_usuario=r["CODIGO_USUARIO"],
            cantidad_gestiones=r["Q_GESTIONES"],
            cantidad_contactos=r["Q_CONTACTOS"],
            cantidad_rut_pagos=r["Q_RUT_PAGOS"],
            monto_pagos=r["MONTO_PAGOS"],
            cantidad_rut_repros=r["Q_RUT_REPROS"],
            monto_repros=r["MONTO_REPROS"],
        )
        for r in rows
    ]


@router.get("/comparativo", response_model=ComparativoResponseCLA)
def comparativo(periodo: str):
    datasets = _ejecutar_multi_resultset(
        "EXEC dbo.SP_Panel1_Comparativo @PeriodoActual=?, @PeriodoAnterior=?",
        (periodo, _periodo_anterior(periodo)),
    )

    def mapear(filas):
        return [
            ComparativoFilaCLA(
                nro_dia=f["NRO_DIA"],
                fecha_actual=f["FECHA_ACTUAL"],
                fecha_anterior=f["FECHA_ANTERIOR"],
                monto_actual=f["MONTO_ACTUAL"],
                monto_anterior=f["MONTO_ANTERIOR"],
                acum_actual=f["ACUM_ACTUAL"],
                acum_anterior=f["ACUM_ANTERIOR"],
                pct_variacion=f["PCT_VARIACION"],
                es_proyeccion=f["ES_PROYECCION"],
            )
            for f in filas
        ]

    return ComparativoResponseCLA(
        pagos=mapear(datasets.get("COMPARATIVO_PAGOS", [])),
        repros=mapear(datasets.get("COMPARATIVO_REPROS", [])),
    )


@router.get("/productividad", response_model=ProductividadResponseCLA)
def productividad(periodo: str):
    general_query = text(
        """
        EXEC dbo.SP_Panel1_Productividad @PeriodoActual = :periodo
        """
    )
    avance_query = text(
        """
        EXEC dbo.SP_Panel1_AvanceEtapa @PeriodoActual = :periodo, @CalcularEnVivo = 1
        """
    )
    with engine.connect() as conn:
        general = conn.execute(general_query, {"periodo": periodo}).mappings().all()
        avance = conn.execute(avance_query, {"periodo": periodo}).mappings().all()

    return ProductividadResponseCLA(
        general=[
            ProductividadFilaCLA(
                cartera=r["CARTERA"],
                cantidad_base=r["Q_BASE"],
                saldo_insoluto=r["SALDO_INSOLUTO"],
                pagos_estudio=r["PAGOS_ESTUDIO"],
                repros_estudio=r["REPROS_ESTUDIO"],
                pagos_inbound=r["PAGOS_INBOUND"],
                total_pagos=r["TOTAL_PAGOS"],
                total_repros=r["TOTAL_REPROS"],
            )
            for r in general
        ],
        avance_etapa=[
            AvanceEtapaFilaCLA(
                cartera=r["CARTERA"],
                categoria=r["CATEGORIA"],
                q_prom_3meses=r["Q_PROM_3MESES"],
                saldo_prom_3meses=r["SALDO_PROM_3MESES"],
                q_mes_anterior=r["Q_MES_ANTERIOR"],
                saldo_mes_anterior=r["SALDO_MES_ANTERIOR"],
                q_mes_actual=r["Q_MES_ACTUAL"],
                saldo_mes_actual=r["SALDO_MES_ACTUAL"],
                q_proyectado_cierre=r["Q_PROYECTADO_CIERRE"],
                saldo_proyectado_cierre=r["SALDO_PROYECTADO_CIERRE"],
            )
            for r in avance
        ],
    )


def _agregar_hoja(wb: Workbook, nombre: str, filas: list[dict]):
    ws = wb.create_sheet(title=nombre[:31])
    if not filas:
        ws.append(["Sin datos"])
        return
    columnas = list(filas[0].keys())
    ws.append(columnas)
    for celda in ws[1]:
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for fila in filas:
        ws.append([fila[c] for c in columnas])
    for col in ws.columns:
        letra = col[0].column_letter
        ws.column_dimensions[letra].width = 22
    ws.freeze_panes = "A2"


def _descargar_sabana(periodo: str, tipo: str, tabla_resumen: str, nombre_archivo: str) -> Response:
    resumen_query = text(f"SELECT * FROM dbo.{tabla_resumen} WHERE PERIODO = :periodo")
    with engine.connect() as conn:
        resumen = [dict(r) for r in conn.execute(resumen_query, {"periodo": periodo}).mappings().all()]

    datasets = _ejecutar_multi_resultset(
        "EXEC dbo.SP_Panel1_Sabanas_Caja_Los_Andes @CARTERA=204, @Periodo=?, @Producto=1, @TIPO=?",
        (periodo, tipo),
    )
    sabana = datasets.get("PANEL_1_SABANA", [])

    if not resumen and not sabana:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sin datos")

    wb = Workbook()
    wb.remove(wb.active)
    _agregar_hoja(wb, f"Resumen {tipo.capitalize()}", resumen)
    _agregar_hoja(wb, f"Sábana {tipo.capitalize()}", sabana)

    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


@router.get("/descargar-pagos")
def descargar_pagos(periodo: str):
    return _descargar_sabana(periodo, "PAGOS", "PANEL1_PAGOS", f"Pagos_CajaLosAndes_{periodo}.xlsx")


@router.get("/descargar-repros")
def descargar_repros(periodo: str):
    return _descargar_sabana(periodo, "REPROS", "PANEL1_REPROS", f"Repros_CajaLosAndes_{periodo}.xlsx")
