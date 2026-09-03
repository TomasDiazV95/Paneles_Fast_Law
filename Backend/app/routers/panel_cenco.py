import io

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from app.db.database import engine
from app.routers.auth import get_current_user
from app.schemas.panel_cenco import (
    ComparativoFilaCenco,
    ComparativoResponseCenco,
    ContactabilidadFilaCenco,
    ContactabilidadResponseCenco,
    EjecutivoCenco,
    EstadoCarteraCenco,
    EstadoCyberFilaCenco,
    EstadoCyberResponseCenco,
    EstadoCyberResumenCenco,
    PagosDiarioCenco,
    PagosFilaCenco,
    PagosResponseCenco,
    PeriodoOptionCenco,
    ReprosDiarioCenco,
    ReprosFilaCenco,
    ReprosResponseCenco,
    ResumenContactoCenco,
    SalidaFilaCenco,
)

router = APIRouter(prefix="/panel/cenco", tags=["panel-cenco"], dependencies=[Depends(get_current_user)])


_ESTADOS_CYBER_VALIDOS = {"ACTUALIZADO", "NO ACTUALIZADO"}


def _validar_estado_cyber(estado: str | None) -> str | None:
    """Normaliza y valida el filtro `estado` de Estado Cyber CENCO.

    Acepta None/vacio (sin filtro, trae todos los estados) o uno de los
    valores que reconoce el SP (`ACTUALIZADO` / `NO ACTUALIZADO`). Cualquier
    otro valor se rechaza con 400 para no dejar pasar strings arbitrarios
    hacia el parametro @Estado del SP.
    """
    if estado is None or not estado.strip():
        return None
    estado_normalizado = estado.strip().upper()
    if estado_normalizado not in _ESTADOS_CYBER_VALIDOS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="estado invalido. Valores permitidos: ACTUALIZADO, NO ACTUALIZADO.",
        )
    return estado_normalizado


def _ejecutar_estado_cyber(
    periodo: str, producto: int, estado: str | None
) -> tuple[dict, list[dict]]:
    """Ejecuta SP_Panel_Cenco_Estado_Cyber (dos result sets: resumen y detalle
    completo, sin paginar) usando una conexion raw + cursor.nextset(), ya que
    SQLAlchemy no expone result sets adicionales de un EXEC via `text()`."""
    raw_conn = engine.raw_connection()
    try:
        cursor = raw_conn.cursor()
        cursor.execute(
            "EXEC dbo.SP_Panel_Cenco_Estado_Cyber @Periodo=?, @Producto=?, @Estado=?",
            (periodo, producto, estado),
        )

        columnas_resumen = [c[0] for c in cursor.description] if cursor.description else []
        # OJO: se usa fetchall() (no fetchone()) aunque el resumen sea una unica
        # fila. Con pyodbc + ODBC Driver 17, si el primer result set tiene una
        # sola fila y se consume con un unico fetchone(), el cursor no queda
        # marcado como agotado y cursor.nextset() puede devolver False de forma
        # incorrecta, perdiendo silenciosamente el segundo result set (detalle).
        # fetchall() SI deja el cursor correctamente posicionado para nextset().
        filas_resumen = cursor.fetchall()
        resumen = dict(zip(columnas_resumen, filas_resumen[0])) if filas_resumen else {}

        detalle: list[dict] = []
        if cursor.nextset() and cursor.description:
            columnas_detalle = [c[0] for c in cursor.description]
            detalle = [dict(zip(columnas_detalle, row)) for row in cursor.fetchall()]

        return resumen, detalle
    finally:
        raw_conn.close()


@router.get("/periodos", response_model=list[PeriodoOptionCenco])
def periodos(cartera: str = "427"):
    query = text(
        """
        SELECT PERIODO, COUNT(*) AS causas
        FROM dbo.PANEL_CENCO_ESTADO_CARTERA
        WHERE ID_CARTERA = :cartera
        GROUP BY PERIODO
        ORDER BY PERIODO DESC
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"cartera": cartera}).mappings().all()
    return [PeriodoOptionCenco(periodo=r["PERIODO"], causas=r["causas"]) for r in rows]


@router.get("/estado-cartera", response_model=list[EstadoCarteraCenco])
def estado_cartera(periodo: str, cartera: str = "427"):
    query = text(
        """
        SELECT ISNULL(CLASIFICACION, 'SIN CLASIFICAR') AS CLASIFICACION,
               ISNULL(Q_CAUSAS, 0) AS Q_CAUSAS,
               ISNULL(CUANTIA_TOTAL, 0) AS CUANTIA_TOTAL,
               ISNULL(TICKET_PROMEDIO, 0) AS TICKET_PROMEDIO
        FROM dbo.PANEL_CENCO_ESTADO_CARTERA
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY Q_CAUSAS DESC
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().all()

    return [
        EstadoCarteraCenco(
            clasificacion=r["CLASIFICACION"],
            cantidad_causas=r["Q_CAUSAS"],
            cuantia_total=r["CUANTIA_TOTAL"],
            ticket_promedio=r["TICKET_PROMEDIO"],
        )
        for r in rows
    ]


@router.get("/contactabilidad", response_model=ContactabilidadResponseCenco)
def contactabilidad(periodo: str, cartera: str = "427"):
    matriz_query = text(
        """
        SELECT ISNULL(CLASIFICACION, 'SIN CLASIFICAR') AS CLASIFICACION,
               ISNULL(TIPO_CONTACTO, 'SIN CONTACTO') AS TIPO_CONTACTO,
               ISNULL(Q_DEUDORES, 0) AS Q_DEUDORES
        FROM dbo.PANEL_CENCO_CONTACTABILIDAD
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY CLASIFICACION, TIPO_CONTACTO
        """
    )
    resumen_query = text(
        """
        SELECT ISNULL(TOTAL_DEUDORES, 0) AS TOTAL_DEUDORES,
               ISNULL(CON_CONTACTO, 0) AS CON_CONTACTO,
               ISNULL(SIN_GESTION, 0) AS SIN_GESTION
        FROM dbo.PANEL_CENCO_RESUMEN_CONTACTO
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        """
    )
    with engine.connect() as conn:
        matriz = conn.execute(matriz_query, {"periodo": periodo, "cartera": cartera}).mappings().all()
        resumen = conn.execute(resumen_query, {"periodo": periodo, "cartera": cartera}).mappings().all()

    return ContactabilidadResponseCenco(
        matriz=[
            ContactabilidadFilaCenco(
                clasificacion=r["CLASIFICACION"],
                tipo_contacto=r["TIPO_CONTACTO"],
                cantidad_deudores=r["Q_DEUDORES"],
            )
            for r in matriz
        ],
        resumen=[
            ResumenContactoCenco(
                total_deudores=r["TOTAL_DEUDORES"],
                con_contacto=r["CON_CONTACTO"],
                sin_gestion=r["SIN_GESTION"],
            )
            for r in resumen
        ],
    )


@router.get("/pagos", response_model=PagosResponseCenco)
def pagos(periodo: str, cartera: str = "427"):
    resumen_query = text(
        """
        SELECT ISNULL(CLASIFICACION, 'SIN CLASIFICAR') AS CLASIFICACION,
               ISNULL(Q_DOCUMENTOS, 0) AS Q_DOCUMENTOS,
               ISNULL(MONTO_TOTAL, 0) AS MONTO_TOTAL
        FROM dbo.PANEL_CENCO_PAGOS
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY CASE WHEN CLASIFICACION = 'TOTAL' THEN 1 ELSE 0 END, MONTO_TOTAL DESC
        """
    )
    diario_query = text(
        """
        SELECT FECHA, ISNULL(MONTO_DIA, 0) AS MONTO_DIA, ISNULL(MONTO_ACUMULADO, 0) AS MONTO_ACUMULADO
        FROM dbo.PANEL_CENCO_PAGOS_DIARIO
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY NRO_DIA
        """
    )
    with engine.connect() as conn:
        resumen = conn.execute(resumen_query, {"periodo": periodo, "cartera": cartera}).mappings().all()
        diario = conn.execute(diario_query, {"periodo": periodo, "cartera": cartera}).mappings().all()

    return PagosResponseCenco(
        resumen=[
            PagosFilaCenco(
                clasificacion=r["CLASIFICACION"],
                cantidad_documentos=r["Q_DOCUMENTOS"],
                monto_total=r["MONTO_TOTAL"],
            )
            for r in resumen
        ],
        diario=[
            PagosDiarioCenco(fecha=r["FECHA"], monto_dia=r["MONTO_DIA"], monto_acumulado=r["MONTO_ACUMULADO"])
            for r in diario
        ],
    )


@router.get("/repros", response_model=ReprosResponseCenco)
def repros(periodo: str, cartera: str = "427"):
    resumen_query = text(
        """
        SELECT ISNULL(CLASIFICACION, 'SIN CLASIFICAR') AS CLASIFICACION,
               ISNULL(Q_DOCUMENTOS, 0) AS Q_DOCUMENTOS,
               ISNULL(MONTO_TOTAL, 0) AS MONTO_TOTAL
        FROM dbo.PANEL_CENCO_REPROS
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY CASE WHEN CLASIFICACION = 'TOTAL' THEN 1 ELSE 0 END, MONTO_TOTAL DESC
        """
    )
    diario_query = text(
        """
        SELECT FECHA, ISNULL(MONTO_DIA, 0) AS MONTO_DIA, ISNULL(MONTO_ACUMULADO, 0) AS MONTO_ACUMULADO
        FROM dbo.PANEL_CENCO_REPROS_DIARIO
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY NRO_DIA
        """
    )
    with engine.connect() as conn:
        resumen = conn.execute(resumen_query, {"periodo": periodo, "cartera": cartera}).mappings().all()
        diario = conn.execute(diario_query, {"periodo": periodo, "cartera": cartera}).mappings().all()

    return ReprosResponseCenco(
        resumen=[
            ReprosFilaCenco(
                clasificacion=r["CLASIFICACION"],
                cantidad_documentos=r["Q_DOCUMENTOS"],
                monto_total=r["MONTO_TOTAL"],
            )
            for r in resumen
        ],
        diario=[
            ReprosDiarioCenco(fecha=r["FECHA"], monto_dia=r["MONTO_DIA"], monto_acumulado=r["MONTO_ACUMULADO"])
            for r in diario
        ],
    )


@router.get("/ejecutivos", response_model=list[EjecutivoCenco])
def ejecutivos(periodo: str, cartera: str = "427"):
    query = text(
        """
        SELECT ISNULL(CODIGO_USUARIO, 'SIN CODIGO') AS CODIGO_USUARIO,
               ISNULL(Q_CONTACTOS, 0) AS Q_CONTACTOS,
               ISNULL(Q_PAGOS, 0) AS Q_PAGOS,
               ISNULL(MONTO_PAGOS, 0) AS MONTO_PAGOS,
               ISNULL(Q_REPROS, 0) AS Q_REPROS,
               ISNULL(MONTO_REPROS, 0) AS MONTO_REPROS
        FROM dbo.PANEL_CENCO_EJECUTIVOS
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY MONTO_PAGOS DESC
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().all()

    return [
        EjecutivoCenco(
            codigo_usuario=r["CODIGO_USUARIO"],
            cantidad_contactos=r["Q_CONTACTOS"],
            cantidad_pagos=r["Q_PAGOS"],
            monto_pagos=r["MONTO_PAGOS"],
            cantidad_repros=r["Q_REPROS"],
            monto_repros=r["MONTO_REPROS"],
        )
        for r in rows
    ]


def _comparativo_filas(conn, tabla: str, periodo: str, cartera: str) -> list[ComparativoFilaCenco]:
    query = text(
        f"""
        SELECT ORD, NRO_DIA, FECHA_ACTUAL, FECHA_ANTERIOR, MONTO_ACTUAL, MONTO_ANTERIOR,
               ACUM_ACTUAL, ACUM_ANTERIOR, TIPO_FILA, GESTOR_ORIGEN
        FROM dbo.{tabla}
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        ORDER BY GESTOR_ORIGEN, ORD, NRO_DIA
        """
    )
    rows = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().all()
    return [
        ComparativoFilaCenco(
            ord=r["ORD"],
            nro_dia=r["NRO_DIA"],
            fecha_actual=r["FECHA_ACTUAL"],
            fecha_anterior=r["FECHA_ANTERIOR"],
            monto_actual=r["MONTO_ACTUAL"],
            monto_anterior=r["MONTO_ANTERIOR"],
            acum_actual=r["ACUM_ACTUAL"],
            acum_anterior=r["ACUM_ANTERIOR"],
            tipo_fila=r["TIPO_FILA"],
            gestor_origen=r["GESTOR_ORIGEN"],
        )
        for r in rows
    ]


@router.get("/comparativo", response_model=ComparativoResponseCenco)
def comparativo(periodo: str, cartera: str = "427"):
    with engine.connect() as conn:
        pagos = _comparativo_filas(conn, "PANEL_CENCO_COMPARATIVO", periodo, cartera)
        repros = _comparativo_filas(conn, "PANEL_CENCO_COMPARATIVO_REPROS", periodo, cartera)

    return ComparativoResponseCenco(pagos=pagos, repros=repros)


_SALIDAS_SP = text("EXEC dbo.SP_Panel_Cenco_Salidas @Periodo = :periodo")


@router.get("/salidas", response_model=list[SalidaFilaCenco])
def salidas(periodo: str):
    with engine.connect() as conn:
        rows = conn.execute(_SALIDAS_SP, {"periodo": int(periodo)}).mappings().all()

    return [
        SalidaFilaCenco(
            cuenta=r["CUENTA"],
            rut=r["RUT"],
            operacion=r["OPERACION"],
            numero_juicio=r["NUMERO_JUICIO"],
            marca_glosa_abogados=r["MARCA_GLOSA_ABOGADOS"],
            marca=r["MARCA"],
            es_duplicado=bool(r["ES_DUPLICADO"]),
            fecha_salida=r["FECHA_SALIDA"],
            casos_dia=r["CASOS_DIA"],
        )
        for r in rows
    ]


@router.get("/salidas/descarga")
def salidas_descarga(periodo: str):
    with engine.connect() as conn:
        rows = conn.execute(_SALIDAS_SP, {"periodo": int(periodo)}).mappings().all()

    if not rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sin datos")

    wb = Workbook()
    ws = wb.active
    ws.title = "Salidas"
    ws.append(["NUMERO_JUICIO", "MARCA"])
    for celda in ws[1]:
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for r in rows:
        ws.append([r["NUMERO_JUICIO"], r["MARCA"]])
    for col in ws.columns:
        letra = col[0].column_letter
        ws.column_dimensions[letra].width = 22
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Salidas_CENCO_{periodo}.xlsx"},
    )


@router.get("/estado-cyber", response_model=EstadoCyberResponseCenco)
def estado_cyber(
    periodo: str,
    producto: int = 5,
    estado: str | None = None,
):
    estado = _validar_estado_cyber(estado)

    resumen, detalle = _ejecutar_estado_cyber(periodo, producto, estado)

    return EstadoCyberResponseCenco(
        resumen=EstadoCyberResumenCenco(
            q_actualizado=resumen.get("Q_ACTUALIZADO", 0),
            q_no_actualizado=resumen.get("Q_NO_ACTUALIZADO", 0),
            q_total=resumen.get("Q_TOTAL", 0),
        ),
        filas=[
            EstadoCyberFilaCenco(
                rut=r["RUT"],
                dv=r["DV"],
                u6id=r["U6ID"],
                operacion=r["OPERACION"],
                tipo_de_cuenta=r["TIPO_DE_CUENTA"],
                rsp_auto_ges=r["RSP_AUTO_GES"],
                resp_jfastco=r["RESP_JFASTCO"],
                estado=r["ESTADO"],
            )
            for r in detalle
        ],
    )


@router.get("/estado-cyber/descarga")
def estado_cyber_descarga(periodo: str, producto: int = 5, estado: str | None = None):
    estado = _validar_estado_cyber(estado)
    _, detalle = _ejecutar_estado_cyber(periodo, producto, estado)

    if not detalle:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sin datos")

    columnas = ["RUT", "DV", "U6ID", "OPERACION", "TIPO_DE_CUENTA", "RSP_AUTO_GES", "RESP_JFASTCO", "ESTADO"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado Cyber"
    ws.append(columnas)
    for celda in ws[1]:
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for r in detalle:
        ws.append([r[c] for c in columnas])
    for col in ws.columns:
        letra = col[0].column_letter
        ws.column_dimensions[letra].width = 22
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=EstadoCyber_CENCO_{periodo}.xlsx"},
    )
