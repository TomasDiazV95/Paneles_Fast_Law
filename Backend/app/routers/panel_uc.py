import io
import re
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from app.db.database import engine
from app.routers.auth import get_current_user
from app.schemas.panel_uc import (
    ActividadDiariaFilaUC,
    CuentaDetalleUC,
    DetalleResponseUC,
    DimensionesResponseUC,
    DimensionFilaUC,
    EmbudoEtapaUC,
    EstadoCarteraFilaUC,
    EvolucionFilaUC,
    FranjaHorariaFilaUC,
    KpiPeriodoUC,
    KpiResumenUC,
    PeriodoOptionUC,
)

router = APIRouter(prefix="/panel/uc", tags=["panel-uc"], dependencies=[Depends(get_current_user)])

CONTACTO_TIPOS = ("CONTACTO TITULAR", "CONTACTO TERCERO")

BUCKET_LABELS = {
    "SIN_GESTION": "Sin gestión",
    "SIN_CONTACTO": "Sin contacto",
    "CONT_SIN_COMP": "Contactado sin compromiso",
    "COMP_PAGO": "Compromiso de pago",
    "COMP_ROTO": "Compromiso roto",
}
BUCKET_ORDER = ["SIN_GESTION", "SIN_CONTACTO", "CONT_SIN_COMP", "COMP_PAGO", "COMP_ROTO"]

ORDEN_COLUMNAS = {
    "rut_deudor": "RUT_DEUDOR",
    "nombre_deudor": "NOMBRE_DEUDOR",
    "monto_documento": "MONTO_DOCUMENTO",
    "saldo_insoluto": "SALDO_INSOLUTO",
    "cantidad_gestiones": "CANTIDAD_GESTIONES",
    "fecha_ultima_gestion": "FECHA_ULTIMA_GESTION",
    "ejecutivo": "EJECUTIVO",
    "bucket": "BUCKET",
    "monto_pagado_periodo": "MONTO_PAGADO_PERIODO",
}


_PERIODO_RE = re.compile(r"^\d{6}$")


def _periodo_valido(periodo: str) -> str:
    """Dependencia reutilizable: valida formato YYYYMM antes de tocar la BD.

    Sin esto, un periodo malformado (ej. 'abcdef') llega íntegro a
    _periodo_anterior() y explota con ValueError sin capturar -> 500 genérico.
    Con la validación acá se responde 400, consistente con el resto de la API.
    """
    if not _PERIODO_RE.match(periodo):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Periodo inválido: formato esperado YYYYMM",
        )
    return periodo


def _periodo_anterior(periodo: str) -> str:
    anio, mes = int(periodo[:4]), int(periodo[4:])
    if mes == 1:
        return f"{anio - 1}12"
    return f"{anio}{mes - 1:02d}"


def _valor(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _kpi_periodo(conn, periodo: str, cartera: int) -> KpiPeriodoUC | None:
    query = text(
        """
        SELECT
            COUNT(*) AS cuentas,
            ISNULL(SUM(MONTO_DOCUMENTO), 0) AS deuda,
            SUM(CASE WHEN BUCKET = 'SIN_GESTION' THEN 1 ELSE 0 END) AS sin_gestion,
            ISNULL(SUM(CASE WHEN BUCKET = 'SIN_GESTION' THEN MONTO_DOCUMENTO ELSE 0 END), 0) AS deuda_sin_gestion,
            ISNULL(SUM(CANTIDAD_GESTIONES), 0) AS gestiones,
            SUM(CASE WHEN TIPO_CONTACTO IN ('CONTACTO TITULAR', 'CONTACTO TERCERO') THEN 1 ELSE 0 END) AS contactos,
            SUM(CASE WHEN TIPO_CONTACTO = 'CONTACTO TITULAR' THEN 1 ELSE 0 END) AS contacto_directo,
            SUM(CASE WHEN BUCKET IN ('COMP_PAGO', 'COMP_ROTO') THEN 1 ELSE 0 END) AS compromisos,
            SUM(CASE WHEN BUCKET = 'COMP_ROTO' THEN 1 ELSE 0 END) AS compromisos_rotos,
            ISNULL(AVG(CAST(CANTIDAD_GESTIONES AS FLOAT)), 0) AS intensidad_media
        FROM dbo.PANEL_UC_CUENTA
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        """
    )
    r = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().first()
    if r is None or r["cuentas"] == 0:
        return None

    cuentas = r["cuentas"]
    gestionadas = cuentas - r["sin_gestion"]
    contactos = r["contactos"] or 0
    contacto_directo = r["contacto_directo"] or 0
    compromisos = r["compromisos"] or 0
    compromisos_rotos = r["compromisos_rotos"] or 0

    return KpiPeriodoUC(
        periodo=periodo,
        cuentas=cuentas,
        deuda=int(r["deuda"]),
        ticket_promedio=(r["deuda"] / cuentas) if cuentas else 0,
        sin_gestion=r["sin_gestion"] or 0,
        deuda_sin_gestion=int(r["deuda_sin_gestion"]),
        cobertura_pct=(gestionadas / cuentas * 100) if cuentas else 0,
        gestiones=r["gestiones"],
        gestiones_por_cuenta_gestionada=(r["gestiones"] / gestionadas) if gestionadas else 0,
        contactos=contactos,
        contactabilidad_pct=(contactos / gestionadas * 100) if gestionadas else 0,
        contacto_directo=contacto_directo,
        contacto_directo_pct=(contacto_directo / contactos * 100) if contactos else 0,
        compromisos=compromisos,
        conversion_compromiso_pct=(compromisos / contacto_directo * 100) if contacto_directo else 0,
        compromisos_rotos=compromisos_rotos,
        incumplimiento_pct=(compromisos_rotos / compromisos * 100) if compromisos else 0,
        intensidad_media=r["intensidad_media"] or 0,
    )


@router.get("/periodos", response_model=list[PeriodoOptionUC])
def periodos(cartera: int = 890):
    query = text(
        """
        SELECT PERIODO, COUNT(*) AS cuentas, MAX(FECHA_PROCESO) AS fecha_proceso
        FROM dbo.PANEL_UC_CUENTA
        WHERE ID_CARTERA = :cartera
        GROUP BY PERIODO
        ORDER BY PERIODO DESC
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"cartera": cartera}).mappings().all()

    return [
        PeriodoOptionUC(periodo=r["PERIODO"], cuentas=r["cuentas"], fecha_proceso=_valor(r["fecha_proceso"]))
        for r in rows
    ]


@router.get("/resumen", response_model=KpiResumenUC)
def resumen(periodo: str = Depends(_periodo_valido), cartera: int = 890):
    with engine.connect() as conn:
        actual = _kpi_periodo(conn, periodo, cartera)
        anterior = _kpi_periodo(conn, _periodo_anterior(periodo), cartera)

    if actual is None:
        actual = KpiPeriodoUC(
            periodo=periodo, cuentas=0, deuda=0, ticket_promedio=0, sin_gestion=0, deuda_sin_gestion=0,
            cobertura_pct=0, gestiones=0, gestiones_por_cuenta_gestionada=0, contactos=0, contactabilidad_pct=0,
            contacto_directo=0, contacto_directo_pct=0, compromisos=0, conversion_compromiso_pct=0,
            compromisos_rotos=0, incumplimiento_pct=0, intensidad_media=0,
        )

    return KpiResumenUC(actual=actual, anterior=anterior)


@router.get("/estado-cartera", response_model=list[EstadoCarteraFilaUC])
def estado_cartera(periodo: str = Depends(_periodo_valido), cartera: int = 890):
    query = text(
        """
        SELECT BUCKET, COUNT(*) AS cuentas, ISNULL(SUM(MONTO_DOCUMENTO), 0) AS deuda,
               ISNULL(SUM(CANTIDAD_GESTIONES), 0) AS gestiones
        FROM dbo.PANEL_UC_CUENTA
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        GROUP BY BUCKET
        """
    )
    with engine.connect() as conn:
        rows = {r["BUCKET"]: r for r in conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().all()}

    total = sum(r["cuentas"] for r in rows.values()) or 1
    return [
        EstadoCarteraFilaUC(
            bucket=bucket,
            etiqueta=BUCKET_LABELS[bucket],
            cuentas=rows[bucket]["cuentas"] if bucket in rows else 0,
            deuda=int(rows[bucket]["deuda"]) if bucket in rows else 0,
            gestiones=rows[bucket]["gestiones"] if bucket in rows else 0,
            pct_cuentas=((rows[bucket]["cuentas"] if bucket in rows else 0) / total * 100),
        )
        for bucket in BUCKET_ORDER
    ]


@router.get("/embudo", response_model=list[EmbudoEtapaUC])
def embudo(periodo: str = Depends(_periodo_valido), cartera: int = 890):
    query = text(
        """
        SELECT
            COUNT(*) AS asignadas,
            SUM(CASE WHEN BUCKET <> 'SIN_GESTION' THEN 1 ELSE 0 END) AS gestionadas,
            SUM(CASE WHEN TIPO_CONTACTO IN ('CONTACTO TITULAR', 'CONTACTO TERCERO') THEN 1 ELSE 0 END) AS contactadas,
            SUM(CASE WHEN TIPO_CONTACTO = 'CONTACTO TITULAR' THEN 1 ELSE 0 END) AS contacto_directo,
            SUM(CASE WHEN BUCKET IN ('COMP_PAGO', 'COMP_ROTO') THEN 1 ELSE 0 END) AS con_compromiso,
            SUM(CASE WHEN MONTO_PAGADO_PERIODO > 0 THEN 1 ELSE 0 END) AS compromiso_cumplido
        FROM dbo.PANEL_UC_CUENTA
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        """
    )
    with engine.connect() as conn:
        r = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().first()

    if r is None or not r["asignadas"]:
        return []

    etapas = [
        ("Cuentas asignadas", r["asignadas"]),
        ("Gestionadas", r["gestionadas"]),
        ("Contactadas", r["contactadas"]),
        ("Contacto directo", r["contacto_directo"]),
        ("Con compromiso", r["con_compromiso"]),
        ("Compromiso cumplido (con pago registrado)", r["compromiso_cumplido"]),
    ]
    total = r["asignadas"] or 1
    resultado = []
    anterior = None
    for etapa, valor in etapas:
        pct_conv = (valor / anterior * 100) if anterior else None
        resultado.append(
            EmbudoEtapaUC(etapa=etapa, cuentas=valor, pct_del_total=(valor / total * 100), pct_conversion_etapa=pct_conv)
        )
        anterior = valor
    return resultado


@router.get("/evolucion", response_model=list[EvolucionFilaUC])
def evolucion(cartera: int = 890):
    query = text(
        """
        SELECT PERIODO,
               COUNT(*) AS cuentas,
               ISNULL(SUM(MONTO_DOCUMENTO), 0) AS deuda,
               ISNULL(SUM(CANTIDAD_GESTIONES), 0) AS gestiones,
               SUM(CASE WHEN BUCKET = 'SIN_GESTION' THEN 1 ELSE 0 END) AS sin_gestion,
               SUM(CASE WHEN TIPO_CONTACTO IN ('CONTACTO TITULAR', 'CONTACTO TERCERO') THEN 1 ELSE 0 END) AS contactos,
               SUM(CASE WHEN BUCKET IN ('COMP_PAGO', 'COMP_ROTO') THEN 1 ELSE 0 END) AS compromisos,
               SUM(CASE WHEN BUCKET = 'COMP_ROTO' THEN 1 ELSE 0 END) AS compromisos_rotos
        FROM dbo.PANEL_UC_CUENTA
        WHERE ID_CARTERA = :cartera
        GROUP BY PERIODO
        ORDER BY PERIODO
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"cartera": cartera}).mappings().all()

    resultado = []
    for r in rows:
        gestionadas = r["cuentas"] - (r["sin_gestion"] or 0)
        resultado.append(
            EvolucionFilaUC(
                periodo=r["PERIODO"],
                cuentas=r["cuentas"],
                deuda=int(r["deuda"]),
                gestiones=r["gestiones"],
                sin_gestion=r["sin_gestion"] or 0,
                contactabilidad_pct=((r["contactos"] or 0) / gestionadas * 100) if gestionadas else 0,
                compromisos=r["compromisos"] or 0,
                compromisos_rotos=r["compromisos_rotos"] or 0,
            )
        )
    return resultado


@router.get("/actividad-diaria", response_model=list[ActividadDiariaFilaUC])
def actividad_diaria(periodo: str = Depends(_periodo_valido), cartera: int = 890):
    query = text(
        """
        SELECT CONVERT(VARCHAR(10), FECHA_ULTIMA_GESTION, 23) AS FECHA, BUCKET, COUNT(*) AS CUENTAS
        FROM dbo.PANEL_UC_CUENTA
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera AND FECHA_ULTIMA_GESTION IS NOT NULL
        GROUP BY FECHA_ULTIMA_GESTION, BUCKET
        ORDER BY FECHA_ULTIMA_GESTION
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().all()

    return [ActividadDiariaFilaUC(fecha=r["FECHA"], bucket=r["BUCKET"], cuentas=r["CUENTAS"]) for r in rows]


@router.get("/franja-horaria", response_model=list[FranjaHorariaFilaUC])
def franja_horaria(periodo: str = Depends(_periodo_valido), cartera: int = 890):
    query = text(
        """
        SELECT HORA_GESTION AS HORA, COUNT(*) AS GESTIONES,
               SUM(CASE WHEN TIPO_CONTACTO IN ('CONTACTO TITULAR', 'CONTACTO TERCERO') THEN 1 ELSE 0 END) AS CONTACTOS
        FROM dbo.PANEL_UC_GESTION
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        GROUP BY HORA_GESTION
        ORDER BY HORA_GESTION
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().all()

    return [FranjaHorariaFilaUC(hora=r["HORA"], gestiones=r["GESTIONES"], contactos=r["CONTACTOS"] or 0) for r in rows]


def _dimension_filas(conn, periodo: str, cartera: int, expresion: str) -> list[DimensionFilaUC]:
    query = text(
        f"""
        SELECT {expresion} AS VALOR,
               COUNT(*) AS CUENTAS,
               ISNULL(SUM(MONTO_DOCUMENTO), 0) AS DEUDA,
               SUM(CASE WHEN TIPO_CONTACTO IN ('CONTACTO TITULAR', 'CONTACTO TERCERO') THEN 1 ELSE 0 END) AS CONTACTOS,
               SUM(CASE WHEN BUCKET IN ('COMP_PAGO', 'COMP_ROTO') THEN 1 ELSE 0 END) AS COMPROMISOS,
               SUM(CASE WHEN BUCKET = 'SIN_GESTION' THEN 1 ELSE 0 END) AS SIN_GESTION
        FROM dbo.PANEL_UC_CUENTA
        WHERE PERIODO = :periodo AND ID_CARTERA = :cartera
        GROUP BY {expresion}
        ORDER BY CUENTAS DESC
        """
    )
    rows = conn.execute(query, {"periodo": periodo, "cartera": cartera}).mappings().all()
    return [
        DimensionFilaUC(
            valor=r["VALOR"],
            cuentas=r["CUENTAS"],
            deuda=int(r["DEUDA"]),
            contactos=r["CONTACTOS"] or 0,
            compromisos=r["COMPROMISOS"] or 0,
            sin_gestion=r["SIN_GESTION"] or 0,
        )
        for r in rows
    ]


@router.get("/dimensiones", response_model=DimensionesResponseUC)
def dimensiones(periodo: str = Depends(_periodo_valido), cartera: int = 890):
    expresiones = {
        # ISNULL trunca al ancho del primer argumento si el segundo es más largo,
        # por eso el CAST se hace a VARCHAR(20) y no VARCHAR(10)
        "prioridad": "ISNULL(CAST(PRIORIDAD AS VARCHAR(20)), 'SIN GESTION')",
        "estado_convenio": "ISNULL(ESTADO_CONVENIO, 'SIN INFORMACION')",
        "ejecutivo": "ISNULL(EJECUTIVO, 'SIN GESTION')",
        "tipificacion": "ISNULL(TIPIFICACION, 'SIN GESTION')",
        "intensidad": (
            "CASE WHEN CANTIDAD_GESTIONES = 0 THEN 'Sin gestión' "
            "WHEN CANTIDAD_GESTIONES <= 2 THEN '1 a 2' "
            "WHEN CANTIDAD_GESTIONES <= 5 THEN '3 a 5' "
            "WHEN CANTIDAD_GESTIONES <= 10 THEN '6 a 10' "
            "ELSE '11 o más' END"
        ),
        "bucket": "BUCKET",
    }
    with engine.connect() as conn:
        datos = {clave: _dimension_filas(conn, periodo, cartera, expr) for clave, expr in expresiones.items()}

    for fila in datos["bucket"]:
        fila.valor = BUCKET_LABELS.get(fila.valor, fila.valor)

    return DimensionesResponseUC(**datos)


def _construir_filtro_detalle(periodo: str, cartera: int, bucket, ejecutivo, tipificacion, estado_convenio, fecha):
    condiciones = ["PERIODO = :periodo", "ID_CARTERA = :cartera"]
    params: dict = {"periodo": periodo, "cartera": cartera}
    if bucket:
        condiciones.append("BUCKET = :bucket")
        params["bucket"] = bucket
    if ejecutivo:
        condiciones.append("ISNULL(EJECUTIVO, 'SIN GESTION') = :ejecutivo")
        params["ejecutivo"] = ejecutivo
    if tipificacion:
        condiciones.append("ISNULL(TIPIFICACION, 'SIN GESTION') = :tipificacion")
        params["tipificacion"] = tipificacion
    if estado_convenio:
        condiciones.append("ISNULL(ESTADO_CONVENIO, 'SIN INFORMACION') = :estado_convenio")
        params["estado_convenio"] = estado_convenio
    if fecha:
        condiciones.append("FECHA_ULTIMA_GESTION = :fecha")
        params["fecha"] = fecha
    return " AND ".join(condiciones), params


def _mapear_fila_detalle(r) -> CuentaDetalleUC:
    return CuentaDetalleUC(
        rut_deudor=r["RUT_DEUDOR"],
        dv_deudor=r["DV_DEUDOR"],
        nombre_deudor=r["NOMBRE_DEUDOR"],
        numero_documento=r["NUMERO_DOCUMENTO"],
        monto_documento=r["MONTO_DOCUMENTO"],
        saldo_insoluto=r["SALDO_INSOLUTO"],
        plazo=r["PLAZO"],
        anho_vehiculo=r["ANHO_VEHICULO"],
        categoria_vehiculo=r["CATEGORIA_VEHICULO"],
        estado_convenio=r["ESTADO_CONVENIO"],
        tipo_contacto=r["TIPO_CONTACTO"],
        tipificacion=r["TIPIFICACION"],
        fecha_ultima_gestion=_valor(r["FECHA_ULTIMA_GESTION"]),
        ejecutivo=r["EJECUTIVO"],
        prioridad=r["PRIORIDAD"],
        cantidad_gestiones=r["CANTIDAD_GESTIONES"],
        fecha_agendamiento=_valor(r["FECHA_AGENDAMIENTO"]),
        monto_agendamiento=r["MONTO_AGENDAMIENTO"],
        monto_pagado_periodo=r["MONTO_PAGADO_PERIODO"],
        cuotas_pagadas=r["CUOTAS_PAGADAS"],
        bucket=r["BUCKET"],
    )


@router.get("/detalle", response_model=DetalleResponseUC)
def detalle(
    periodo: str = Depends(_periodo_valido),
    cartera: int = 890,
    bucket: str | None = None,
    ejecutivo: str | None = None,
    tipificacion: str | None = None,
    estado_convenio: str | None = None,
    fecha: str | None = None,
    orden: str = "monto_documento",
    direccion: str = "desc",
    pagina: int = 1,
    tamano_pagina: int = 200,
):
    where_sql, params = _construir_filtro_detalle(periodo, cartera, bucket, ejecutivo, tipificacion, estado_convenio, fecha)
    columna_orden = ORDEN_COLUMNAS.get(orden, "MONTO_DOCUMENTO")
    direccion_sql = "ASC" if direccion.lower() == "asc" else "DESC"
    pagina = max(1, pagina)
    tamano_pagina = min(max(1, tamano_pagina), 500)
    offset = (pagina - 1) * tamano_pagina

    # Este SQL Server no soporta OFFSET/FETCH (compatibilidad ~2008), se pagina
    # con ROW_NUMBER() igual que el resto de los SP existentes en el proyecto.
    fila_desde = offset + 1
    fila_hasta = offset + tamano_pagina

    total_query = text(f"SELECT COUNT(*) FROM dbo.PANEL_UC_CUENTA WHERE {where_sql}")
    filas_query = text(
        f"""
        SELECT RUT_DEUDOR, DV_DEUDOR, NOMBRE_DEUDOR, NUMERO_DOCUMENTO, MONTO_DOCUMENTO, SALDO_INSOLUTO,
               PLAZO, ANHO_VEHICULO, CATEGORIA_VEHICULO, ESTADO_CONVENIO, TIPO_CONTACTO, TIPIFICACION,
               FECHA_ULTIMA_GESTION, EJECUTIVO, PRIORIDAD, CANTIDAD_GESTIONES, FECHA_AGENDAMIENTO,
               MONTO_AGENDAMIENTO, MONTO_PAGADO_PERIODO, CUOTAS_PAGADAS, BUCKET
        FROM (
            SELECT *, ROW_NUMBER() OVER (ORDER BY {columna_orden} {direccion_sql}) AS RN
            FROM dbo.PANEL_UC_CUENTA
            WHERE {where_sql}
        ) X
        WHERE RN BETWEEN :fila_desde AND :fila_hasta
        ORDER BY RN
        """
    )
    with engine.connect() as conn:
        total = conn.execute(total_query, params).scalar() or 0
        filas = conn.execute(
            filas_query, {**params, "fila_desde": fila_desde, "fila_hasta": fila_hasta}
        ).mappings().all()

    return DetalleResponseUC(
        total=total,
        pagina=pagina,
        tamano_pagina=tamano_pagina,
        filas=[_mapear_fila_detalle(r) for r in filas],
    )


@router.get("/descarga")
def descarga(
    periodo: str = Depends(_periodo_valido),
    cartera: int = 890,
    bucket: str | None = None,
    ejecutivo: str | None = None,
    tipificacion: str | None = None,
    estado_convenio: str | None = None,
):
    where_sql, params = _construir_filtro_detalle(periodo, cartera, bucket, ejecutivo, tipificacion, estado_convenio, None)
    query = text(
        f"""
        SELECT RUT_DEUDOR, DV_DEUDOR, NOMBRE_DEUDOR, NUMERO_DOCUMENTO, MONTO_DOCUMENTO, SALDO_INSOLUTO,
               PLAZO, ANHO_VEHICULO, CATEGORIA_VEHICULO, ESTADO_CONVENIO, TIPO_CONTACTO, TIPIFICACION,
               FECHA_ULTIMA_GESTION, EJECUTIVO, PRIORIDAD, CANTIDAD_GESTIONES, FECHA_AGENDAMIENTO,
               MONTO_AGENDAMIENTO, MONTO_PAGADO_PERIODO, CUOTAS_PAGADAS, BUCKET
        FROM dbo.PANEL_UC_CUENTA
        WHERE {where_sql}
        ORDER BY MONTO_DOCUMENTO DESC
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, params).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"UC_{periodo}"
    if not rows:
        ws.append(["Sin datos"])
    else:
        columnas = list(rows[0].keys())
        ws.append(columnas)
        for celda in ws[1]:
            celda.font = Font(color="FFFFFF", bold=True)
            celda.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for fila in rows:
            ws.append([_valor(fila[c]) for c in columnas])
        for col in ws.columns:
            letra = col[0].column_letter
            ws.column_dimensions[letra].width = 20
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=detalle_uc_{periodo}.xlsx"},
    )
