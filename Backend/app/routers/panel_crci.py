"""
Panel CRCI — Seguimientos de Asignaciones judiciales.

Fuente funcional: panel legado Flask (crci_app4.py) que apuntaba al mismo
server+DB (192.168.100.136 / JUDICIAL) que ya usa el `engine` de este
proyecto, así que no se agrega conexión nueva. La tabla `CRCI_SEGUIMIENTOS_
ASIGNACIONES` la llena un job de SSMS que corre fuera de esta app (no se
implementa endpoint de carga/actualización acá).

El SP `dbo.SP_CRCI_METRICAS_ASIGNACIONES` ya existe en la base y NO se
modifica. Su definición fue verificada directamente contra la BD
(OBJECT_DEFINITION) para confirmar los alias reales de columnas que
devuelve: FECHA_PROCESO_CRCI, ID_PRODUCTO, MES, ANIO, TOTAL_REGISTROS,
STOCK, FLUJO_MENSUAL_ASIGNACION, REINGRESOS, FLUJO_MENSUAL_INGRESO,
APERCIBIMIENTO, RETIRA_DEMANDA, MANDAMIENTO. Con eso confirmado, el acceso
al resultset se hace por nombre (`.mappings()`), igual que el resto del
proyecto (panel_cla, panel_uc), en vez de por índice posicional como hacía
el panel legado.
"""

import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from app.db.database import engine
from app.routers.auth import get_current_user
from app.schemas.panel_crci import (
    IteracionesResponseCRCI,
    MetricasCRCI,
    MovimientoDiarioFilaCRCI,
    MovimientoDiarioResponseCRCI,
    ProductoCRCI,
)

router = APIRouter(prefix="/panel/crci", tags=["panel-crci"], dependencies=[Depends(get_current_user)])

_PRODUCTOS_ACTIVOS_QUERY = text(
    "SELECT ID_PRODUCTO, NOMBRE_MANDANTE FROM dbo.TBL_CRCI_PRODUCTOS WHERE ACTIVO = 1 ORDER BY NOMBRE_MANDANTE"
)

_PRODUCTO_NOMBRE_QUERY = text(
    "SELECT NOMBRE_MANDANTE FROM dbo.TBL_CRCI_PRODUCTOS WHERE ID_PRODUCTO = :id_producto"
)

_PRODUCTO_ACTIVO_QUERY = text(
    "SELECT 1 FROM dbo.TBL_CRCI_PRODUCTOS WHERE ID_PRODUCTO = :id_producto AND ACTIVO = 1"
)


def _productos_activos(conn) -> dict[int, str]:
    rows = conn.execute(_PRODUCTOS_ACTIVOS_QUERY).mappings().all()
    return {r["ID_PRODUCTO"]: r["NOMBRE_MANDANTE"] for r in rows}


def _nombre_producto(conn, id_producto: int) -> str:
    row = conn.execute(_PRODUCTO_NOMBRE_QUERY, {"id_producto": id_producto}).first()
    return row[0] if row else "PRODUCTO"


_METRICAS_SP = text(
    """
    EXEC dbo.SP_CRCI_METRICAS_ASIGNACIONES
        @ID_PRODUCTO        = :id_producto,
        @MES                = :mes,
        @ANIO               = :anio,
        @FECHA_PROCESO_CRCI = :fecha_proceso
    """
)


def _producto_valido(id_producto: int) -> int:
    with engine.connect() as conn:
        existe = conn.execute(_PRODUCTO_ACTIVO_QUERY, {"id_producto": id_producto}).first()
    if existe is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="id_producto no reconocido o inactivo")
    return id_producto


def _valor(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _ejecutar_metricas(conn, id_producto: int, mes: int, anio: int, fecha_proceso: str):
    return conn.execute(
        _METRICAS_SP,
        {"id_producto": id_producto, "mes": mes, "anio": anio, "fecha_proceso": fecha_proceso},
    ).mappings().first()


def _fila_metricas(row) -> MetricasCRCI:
    return MetricasCRCI(
        fecha_proceso=str(row["FECHA_PROCESO_CRCI"]),
        id_producto=row["ID_PRODUCTO"],
        mes=row["MES"],
        anio=row["ANIO"],
        total=row["TOTAL_REGISTROS"] or 0,
        stock=row["STOCK"] or 0,
        flujo_asignacion=row["FLUJO_MENSUAL_ASIGNACION"] or 0,
        reingresos=row["REINGRESOS"] or 0,
        flujo_ingreso=row["FLUJO_MENSUAL_INGRESO"] or 0,
        apercibimiento=row["APERCIBIMIENTO"] or 0,
        retira_demanda=row["RETIRA_DEMANDA"] or 0,
        mandamiento=row["MANDAMIENTO"] or 0,
    )


# --- Clasificación "METRICA" de la sábana (descarga) ----------------------
# Reimplementación fiel de calcular_metrica() del panel legado (pandas), pero
# sin pandas: acá las columnas llegan como None (no NaN) porque se leen vía
# SQLAlchemy/pyodbc directo, así que no existen los artefactos 'nan'/'None'
# como string que el código legado tenía que filtrar por venir de pandas.

_ETAPAS_APERCIBIMIENTO = {
    "PENDIENTE ACOMPAÑADA DOCUMENTOS",
    "INGRESO DEMANDA A TRIBUNAL – REINGRESO",
    "INGRESO DEMANDA A TRIBUNAL",
    "NO DA CURSO DE A LA DEMANDA",
    "POSPONE INICIO DE TRAMITACIÓN",
    "",
    "None",
    "nan",
}

_ETAPAS_EXCLUIR_MANDAMIENTO = {
    "NO DA CURSO DE A LA DEMANDA",
    "DOCUMENTOS RETIRADO DE TRIBUNALES Y CUSTODIADO EN OFICINA",
    "PENDIENTE ACOMPAÑADA DOCUMENTOS",
    "INGRESO DEMANDA A TRIBUNAL – REINGRESO",
    "INGRESO DEMANDA A TRIBUNAL",
    "ASUME PATROCINIO - CAUSA REASIGNADA",
    "POSPONE INICIO DE TRAMITACIÓN",
    "0",
    "",
    "None",
    "nan",
}


def _texto(valor) -> str:
    """Replica `str(valor or '').strip()` del legado: None, '' y 0 colapsan a
    cadena vacía (0 en ROL se usa como "sin rol asignado")."""
    if not valor:
        return ""
    return str(valor).strip()


def _fecha_coincide(fecha, mes: int, anio: int) -> bool:
    if fecha is None:
        return False
    if isinstance(fecha, str):
        try:
            fecha = datetime.fromisoformat(fecha)
        except ValueError:
            return False
    if isinstance(fecha, (datetime, date)):
        return fecha.month == mes and fecha.year == anio
    return False


def _vacio(valor) -> bool:
    return valor is None or (isinstance(valor, str) and valor.strip() == "")


def _calcular_metrica(row: dict, mes: int, anio: int) -> str:
    metricas: list[str] = []
    rol = _texto(row.get("ROL"))
    etapa = _texto(row.get("CLASIFICACION_ETAPAS"))
    mes_ok = _fecha_coincide(row.get("FECHA_ASIGNACION"), mes, anio)
    sin_custodia = _vacio(row.get("FECHA_CUSTODIA"))

    if rol in ("", "0", "None", "nan"):
        metricas.append("STOCK")
    if rol not in ("", "0", "None", "nan") and mes_ok:
        metricas.append("FLUJO MENSUAL")
    if (
        etapa == "NO DA CURSO DE A LA DEMANDA"
        or "RETIRO DE DEMANDA Y DEVOLUCION DE DOCUMENTOS" in etapa
        or etapa == "DOCUMENTOS RETIRADO DE TRIBUNALES Y CUSTODIADO EN OFICINA"
        or etapa == "PENDIENTE ACOMPAÑADA DOCUMENTOS"
    ):
        metricas.append("REINGRESO")
    if mes_ok and sin_custodia and (etapa in _ETAPAS_APERCIBIMIENTO or etapa.startswith("CUMPLE LO ORDENADO")):
        metricas.append("APERCIBIMIENTO")
    if (
        etapa == "NO DA CURSO DE A LA DEMANDA"
        or "RETIRO DE DEMANDA Y DEVOLUCION DE DOCUMENTOS" in etapa
        or etapa == "DOCUMENTOS RETIRADO DE TRIBUNALES Y CUSTODIADO EN OFICINA"
    ):
        metricas.append("RETIRA DEMANDA")
    if (
        etapa not in _ETAPAS_EXCLUIR_MANDAMIENTO
        and "RETIRO DE DEMANDA Y DEVOLUCION DE DOCUMENTOS" not in etapa
        and not etapa.startswith("CUMPLE LO ORDENADO")
        and etapa != ""
    ):
        metricas.append("MANDAMIENTO")
    return " | ".join(metricas) if metricas else "SIN CLASIFICAR"


@router.get("/productos", response_model=list[ProductoCRCI])
def productos():
    with engine.connect() as conn:
        activos = _productos_activos(conn)
    return [ProductoCRCI(id_producto=k, nombre=v) for k, v in activos.items()]


@router.get("/iteraciones", response_model=IteracionesResponseCRCI)
def iteraciones(
    id_producto: int = Depends(_producto_valido),
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
):
    prefijo = f"{anio}{mes:02d}%"
    query = text(
        """
        SELECT DISTINCT FECHA_PROCESO_CRCI
        FROM dbo.CRCI_SEGUIMIENTOS_ASIGNACIONES
        WHERE ID_PRODUCTO = :id_producto AND FECHA_PROCESO_CRCI LIKE :prefijo
        ORDER BY FECHA_PROCESO_CRCI
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"id_producto": id_producto, "prefijo": prefijo}).fetchall()
    return IteracionesResponseCRCI(iteraciones=[r[0] for r in rows])


@router.get("/metricas", response_model=MetricasCRCI)
def metricas(
    id_producto: int = Depends(_producto_valido),
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    fecha_proceso: str = Query(...),
):
    with engine.connect() as conn:
        row = _ejecutar_metricas(conn, id_producto, mes, anio, fecha_proceso)
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sin resultados para esos parámetros")
    return _fila_metricas(row)


@router.get("/movimiento-diario", response_model=MovimientoDiarioResponseCRCI)
def movimiento_diario(
    id_producto: int = Depends(_producto_valido),
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
):
    prefijo = f"{anio}{mes:02d}%"
    iter_query = text(
        """
        SELECT DISTINCT FECHA_PROCESO_CRCI
        FROM dbo.CRCI_SEGUIMIENTOS_ASIGNACIONES
        WHERE ID_PRODUCTO = :id_producto AND FECHA_PROCESO_CRCI LIKE :prefijo
        ORDER BY FECHA_PROCESO_CRCI
        """
    )
    with engine.connect() as conn:
        todas = [r[0] for r in conn.execute(iter_query, {"id_producto": id_producto, "prefijo": prefijo}).fetchall()]
        if not todas:
            return MovimientoDiarioResponseCRCI(dias=[])

        # El prefijo de 8 caracteres es el día (YYYYMMDD); si hay varias
        # corridas en el mismo día, se recalcula con la última (igual que el
        # panel legado: el dict sobreescribe en orden ascendente).
        dias_map: dict[str, str] = {}
        for it in todas:
            dias_map[it[:8]] = it

        # Se ejecuta el SP una vez por día (máx. ~31 llamadas/mes, acotado por
        # el propio negocio: nunca hay más iteraciones que días del mes). Se
        # evaluó reemplazar este loop por una sola query agregada por día
        # (como hace panel_uc en /actividad-diaria), pero SP_CRCI_METRICAS_
        # ASIGNACIONES no es un simple agregado: aplica reglas de negocio
        # (parseo de CLASIFICACION_ETAPAS, ventanas de fecha, exclusiones)
        # verificadas contra su definición real en la BD. Reimplementar esa
        # lógica en SQL fuera del SP la duplicaría en dos lugares (mismo
        # riesgo ya detectado en el proyecto con SP3 duplicando a SP1 en CLA),
        # así que se mantiene el loop reutilizando la misma conexión (a
        # diferencia del panel Flask legado, que abría una conexión nueva por
        # día). Con el volumen esperado (un mes calendario) el costo es
        # aceptable.
        resultados = []
        for dia_base, fecha_proceso in sorted(dias_map.items()):
            row = _ejecutar_metricas(conn, id_producto, mes, anio, fecha_proceso)
            if row is None:
                continue
            fila = _fila_metricas(row)
            resultados.append(
                MovimientoDiarioFilaCRCI(
                    dia=f"{dia_base[6:8]}/{dia_base[4:6]}/{dia_base[0:4]}",
                    fecha_proceso=fecha_proceso,
                    total=fila.total,
                    stock=fila.stock,
                    flujo_asignacion=fila.flujo_asignacion,
                    reingresos=fila.reingresos,
                    flujo_ingreso=fila.flujo_ingreso,
                    apercibimiento=fila.apercibimiento,
                    retira_demanda=fila.retira_demanda,
                    mandamiento=fila.mandamiento,
                )
            )
    return MovimientoDiarioResponseCRCI(dias=resultados)


@router.get("/descarga")
def descarga(
    id_producto: int = Depends(_producto_valido),
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000, le=2100),
    fecha_proceso: str = Query(...),
):
    # SELECT * intencional: esta es la "sábana" completa de la tabla origen
    # (todas sus columnas), igual que el panel legado — no está modelada en
    # un schema porque el set de columnas lo define el job de SSMS que carga
    # la tabla, no esta API.
    query = text(
        """
        SELECT *
        FROM dbo.CRCI_SEGUIMIENTOS_ASIGNACIONES
        WHERE ID_PRODUCTO = :id_producto AND FECHA_PROCESO_CRCI = :fecha_proceso
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(query, {"id_producto": id_producto, "fecha_proceso": fecha_proceso}).mappings().all()
        nombre_producto = _nombre_producto(conn, id_producto).replace(" ", "_")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sabana"

    if not rows:
        ws.append(["Sin datos"])
    else:
        columnas_originales = list(rows[0].keys())
        columnas = columnas_originales + ["METRICA"]
        ws.append(columnas)
        # Mismo estilo de header que panel_cla/panel_uc para consistencia
        # visual entre las descargas de la app (PatternFill+Font sólido azul).
        for celda in ws[1]:
            celda.font = Font(color="FFFFFF", bold=True)
            celda.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for fila in rows:
            data = dict(fila)
            metrica = _calcular_metrica(data, mes, anio)
            ws.append([_valor(data[c]) for c in columnas_originales] + [metrica])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Sabana_{nombre_producto}_{fecha_proceso}.xlsx"},
    )
